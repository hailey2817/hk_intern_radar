#!/usr/bin/env python3
"""Local server for the Hong Kong internship tracker."""

from __future__ import annotations

import json
import mimetypes
import os
import re
import threading
from datetime import date, datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WEB_DIR = ROOT / "web"
SEED_DATA_DIR = ROOT / "data"
DATA_DIR = Path(os.environ.get("RUNTIME_DATA_DIR", SEED_DATA_DIR)).resolve()
WORKBOOK = ROOT / "Hong_Kong_2027_Internship_Tracker.xlsx"
STATE_FILE = DATA_DIR / "user_state.json"
CUSTOM_FILE = DATA_DIR / "custom_opportunities.json"
RESEARCH_FILE = DATA_DIR / "research_results.json"
RESEARCH_STATUS_FILE = DATA_DIR / "research_status.json"
RESEARCH_REQUEST_FILE = DATA_DIR / "research_request.json"
COMPANY_UNIVERSE_FILE = SEED_DATA_DIR / "company_universe.json"
STATE_LOCK = threading.Lock()


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return default


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    temporary.replace(path)


def iso_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def load_workbook_opportunities() -> list[dict[str, Any]]:
    if not WORKBOOK.exists():
        return []
    wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
    ws = wb["Opportunities"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    key_map = {
        "ID": "id",
        "Category": "category",
        "Company": "company",
        "Program / Role": "role",
        "Season": "season",
        "Location": "location",
        "Application Status": "applicationStatus",
        "Opening Date": "openingDate",
        "Closing Date": "closingDate",
        "Deadline Time Zone": "deadlineTimezone",
        "Rolling?": "rolling",
        "Placement Timeline": "placementTimeline",
        "Deadline Alert": "workbookAlert",
        "Next Action": "nextAction",
        "Job Description": "description",
        "Requirements": "requirements",
        "Language / Work Rights": "languageWorkRights",
        "Recruitment Stages": "recruitmentStages",
        "Document Checklist": "documents",
        "Official Application / Source": "sourceUrl",
        "Date Confidence": "confidence",
        "Notes": "notes",
    }
    items: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        if not raw.get("ID"):
            continue
        item = {key_map[key]: iso_value(value) for key, value in raw.items() if key in key_map}
        item["origin"] = "workbook"
        items.append(item)
    return items


def baseline_research() -> dict[str, Any]:
    workbook_items = load_workbook_opportunities()
    opportunities = [item for item in workbook_items if item.get("applicationStatus") == "Upcoming" or str(item.get("applicationStatus", "")).startswith("Open")]
    watchlist = [item for item in workbook_items if item not in opportunities]
    return {
        "generatedAt": datetime.fromtimestamp(WORKBOOK.stat().st_mtime, timezone.utc).isoformat() if WORKBOOK.exists() else None,
        "cycle": "Winter 2026-27 and Summer 2027",
        "summary": {
            "open": sum(1 for item in opportunities if str(item.get("applicationStatus", "")).startswith("Open")),
            "upcoming": sum(1 for item in opportunities if item.get("applicationStatus") == "Upcoming"),
            "closingSoon": 0,
            "newSinceLastRun": 0,
        },
        "opportunities": opportunities,
        "watchlist": watchlist,
        "changes": [{"type": "baseline", "message": "Initial workbook baseline loaded. The research agent will replace this after its first run."}],
    }


def load_research() -> dict[str, Any]:
    research = read_json(RESEARCH_FILE, None) or baseline_research()
    items = research.get("opportunities", [])
    watchlist = research.get("watchlist", [])
    covered_companies = {str(item.get("company", "")).casefold() for item in items + watchlist}
    for company in read_json(COMPANY_UNIVERSE_FILE, []):
        if str(company.get("company", "")).casefold() not in covered_companies:
            watchlist.append({
                "company": company.get("company"),
                "role": company.get("segment", "Internship programme"),
                "category": company.get("category"),
                "sourceUrl": company.get("sourceUrl"),
                "notes": company.get("reason"),
                "priority": company.get("priority", "Expanded"),
                "dateConfidence": "watchlist-unconfirmed",
            })
    research["watchlist"] = watchlist
    user_state = read_json(STATE_FILE, {})
    for item in items:
        item.setdefault("id", f"agent-{abs(hash((item.get('company'), item.get('role'))))}")
        state = user_state.get(item["id"], {})
        item["pipelineStatus"] = state.get("pipelineStatus", "Interested")
        item["priority"] = state.get("priority", "Medium")
        item["userNotes"] = state.get("userNotes", "")
        item["lastUpdated"] = state.get("lastUpdated")
    research["opportunities"] = items
    return research


class AppHandler(BaseHTTPRequestHandler):
    server_version = "HKInternshipRadar/1.0"

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"[{self.log_date_time_string()}] {fmt % args}")

    def send_json(self, payload: Any, status: int = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=True).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def read_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/opportunities":
            research = load_research()
            request = read_json(RESEARCH_REQUEST_FILE, {"requested": False})
            status = read_json(RESEARCH_STATUS_FILE, {"status": "waiting", "message": "Waiting for the first research-agent run"})
            self.send_json({
                "items": research.get("opportunities", []),
                "watchlist": research.get("watchlist", []),
                "changes": research.get("changes", []),
                "meta": {
                    "count": len(research.get("opportunities", [])),
                    "generatedAt": research.get("generatedAt"),
                    "cycle": research.get("cycle"),
                    "summary": research.get("summary", {}),
                    "researchStatus": status,
                    "refreshRequest": request,
                    "scheduledResearch": "Every Monday at 09:00 Hong Kong time",
                    "onDemandWorker": "Checks queued requests hourly",
                },
            })
            return
        if parsed.path == "/api/health":
            self.send_json({"status": "ok", "time": utc_now(), "research": read_json(RESEARCH_STATUS_FILE, {})})
            return
        if parsed.path == "/api/research/request":
            self.send_json(read_json(RESEARCH_REQUEST_FILE, {"requested": False}))
            return
        self.serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            body = self.read_body()
        except (json.JSONDecodeError, UnicodeDecodeError):
            self.send_json({"error": "Invalid JSON body"}, HTTPStatus.BAD_REQUEST)
            return

        if parsed.path == "/api/refresh":
            request = read_json(RESEARCH_REQUEST_FILE, {})
            if request.get("requested"):
                self.send_json({"accepted": False, "request": request}, HTTPStatus.CONFLICT)
                return
            request = {"requested": True, "requestedAt": utc_now(), "reason": body.get("reason", "User requested refresh from web application")}
            with STATE_LOCK:
                write_json(RESEARCH_REQUEST_FILE, request)
            self.send_json({"accepted": True, "request": request}, HTTPStatus.ACCEPTED)
            return

        if parsed.path == "/api/research/import":
            expected_token = os.environ.get("RESEARCH_SYNC_TOKEN", "")
            supplied_token = self.headers.get("X-Research-Token", "")
            if not expected_token or supplied_token != expected_token:
                self.send_json({"error": "Unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            research = body.get("research")
            if not isinstance(research, dict) or not isinstance(research.get("opportunities"), list):
                self.send_json({"error": "A valid research payload is required"}, HTTPStatus.BAD_REQUEST)
                return
            with STATE_LOCK:
                write_json(RESEARCH_FILE, research)
                write_json(RESEARCH_STATUS_FILE, body.get("status") or {
                    "status": "complete",
                    "startedAt": body.get("startedAt"),
                    "completedAt": utc_now(),
                    "message": f"Imported {len(research.get('opportunities', []))} researched opportunities",
                })
                write_json(RESEARCH_REQUEST_FILE, {"requested": False, "completedAt": utc_now()})
            self.send_json({"imported": True, "opportunities": len(research.get("opportunities", []))})
            return

        if parsed.path == "/api/state":
            opportunity_id = str(body.get("id", "")).strip()
            if not opportunity_id:
                self.send_json({"error": "Opportunity id is required"}, HTTPStatus.BAD_REQUEST)
                return
            with STATE_LOCK:
                state = read_json(STATE_FILE, {})
                current = state.get(opportunity_id, {})
                for field in ("pipelineStatus", "priority", "userNotes"):
                    if field in body:
                        current[field] = body[field]
                current["lastUpdated"] = utc_now()
                state[opportunity_id] = current
                write_json(STATE_FILE, state)
            self.send_json({"saved": True, "state": current})
            return

        if parsed.path == "/api/opportunities":
            required = ("company", "role", "category", "sourceUrl")
            missing = [field for field in required if not str(body.get(field, "")).strip()]
            if missing:
                self.send_json({"error": f"Missing required fields: {', '.join(missing)}"}, HTTPStatus.BAD_REQUEST)
                return
            custom = read_json(CUSTOM_FILE, [])
            max_id = max([int(re.sub(r"\D", "", item.get("id", "0")) or 0) for item in load_research().get("opportunities", [])] + [0])
            item = {
                "id": f"HK27-{max_id + 1:03d}",
                "category": body["category"],
                "company": body["company"].strip(),
                "role": body["role"].strip(),
                "season": body.get("season", "Summer 2027"),
                "location": body.get("location", "Hong Kong"),
                "applicationStatus": body.get("applicationStatus", "Monitor"),
                "openingDate": body.get("openingDate") or None,
                "closingDate": body.get("closingDate") or None,
                "deadlineTimezone": body.get("deadlineTimezone", "Confirm on official page"),
                "rolling": body.get("rolling", "Unknown"),
                "placementTimeline": body.get("placementTimeline", "To be confirmed"),
                "nextAction": body.get("nextAction", "Review official source"),
                "description": body.get("description", ""),
                "requirements": body.get("requirements", ""),
                "languageWorkRights": body.get("languageWorkRights", ""),
                "recruitmentStages": body.get("recruitmentStages", ""),
                "documents": body.get("documents", "CV; transcript if requested; work authorization details"),
                "sourceUrl": body["sourceUrl"].strip(),
                "confidence": body.get("confidence", "User-added; requires verification"),
                "notes": body.get("notes", ""),
                "origin": "custom",
            }
            with STATE_LOCK:
                custom.append(item)
                write_json(CUSTOM_FILE, custom)
            self.send_json({"created": True, "item": item}, HTTPStatus.CREATED)
            return

        self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)

    def serve_static(self, raw_path: str) -> None:
        path = unquote(raw_path)
        if path == "/":
            path = "/index.html"
        candidate = (WEB_DIR / path.lstrip("/")).resolve()
        if WEB_DIR.resolve() not in candidate.parents and candidate != WEB_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        if not candidate.exists() or not candidate.is_file():
            candidate = WEB_DIR / "index.html"
        content = candidate.read_bytes()
        content_type, _ = mimetypes.guess_type(candidate.name)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", f"{content_type or 'application/octet-stream'}; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)


def main() -> None:
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8080"))
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"HK Internship Radar running at http://{host}:{port}")
    print("Reading agent results from data/research_results.json")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
