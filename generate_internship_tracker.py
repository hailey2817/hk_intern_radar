from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path("Hong_Kong_2027_Internship_Tracker.xlsx")
TODAY = date(2026, 7, 15)

NAVY = "17324D"
TEAL = "087E8B"
GOLD = "F2C14E"
LIGHT_BLUE = "DCEAF4"
LIGHT_GREEN = "DDF2E4"
LIGHT_YELLOW = "FFF1C7"
LIGHT_RED = "FADBD8"
WHITE = "FFFFFF"
GRAY = "6B7280"
GRID = "D1D5DB"


def d(value):
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None


opportunities = [
    {
        "category": "Investment Banking & Finance",
        "company": "HSBC",
        "role": "Investment Banking - Internship",
        "season": "Summer 2027",
        "location": "Central, Hong Kong",
        "status": "Open",
        "open_date": d("2026-07-06"),
        "deadline": d("2026-10-30"),
        "deadline_tz": "Hong Kong time (confirm on form)",
        "rolling": "May close early",
        "placement": "2027-06-14; 10 weeks",
        "description": "Support strategic transactions, fundraising and advisory work for global companies.",
        "requirements": "Penultimate-year fit; curiosity, learning agility, teamwork, communication and comfort in a fast-paced environment. Finance degree not required.",
        "language_work_rights": "Role is in Hong Kong; verify form-specific language terms. Applicant has HK work rights.",
        "stages": "Online application; work-based assessments; later interview or assessment centre stages may be virtual or in person.",
        "documents": "CV; transcript if requested; work-authorization details; assessment preparation",
        "source": "https://apply.careers.hsbc.com/emergingtalent/job/Central-Investment-Banking-Internship-Hong/1365768357/",
        "confidence": "Confirmed 2027 official role",
        "notes": "Apply early; HSBC states programmes may close before the published date.",
    },
    {
        "category": "Investment Banking & Finance",
        "company": "HSBC",
        "role": "Investment Banking - Infrastructure Finance Internship",
        "season": "Summer 2027",
        "location": "Central, Hong Kong",
        "status": "Open",
        "open_date": d("2026-07-06"),
        "deadline": d("2026-10-30"),
        "deadline_tz": "Hong Kong time (confirm on form)",
        "rolling": "May close early",
        "placement": "2027-06-14; duration on vacancy",
        "description": "Help deliver financing solutions for major infrastructure projects and the energy transition.",
        "requirements": "Penultimate-year internship profile; analytical ability, collaboration, learning agility and clear communication.",
        "language_work_rights": "Verify vacancy-specific language terms; Hong Kong work rights available.",
        "stages": "Online application and work-based assessments; interview stages as specified in candidate portal.",
        "documents": "CV; transcript if requested; work-authorization details; assessment preparation",
        "source": "https://www.hsbc.com/careers/students-and-graduates/find-a-programme?programme-type=internship-programme",
        "confidence": "Confirmed 2027 official listing",
        "notes": "Use HSBC programme filter for Hong Kong SAR and Investment Banking.",
    },
    {
        "category": "Investment Banking & Finance",
        "company": "HSBC",
        "role": "Markets - Sales and Trading Internship",
        "season": "Summer 2027",
        "location": "Central, Hong Kong",
        "status": "Open",
        "open_date": d("2026-07-06"),
        "deadline": d("2026-10-30"),
        "deadline_tz": "Hong Kong time (confirm on form)",
        "rolling": "May close early",
        "placement": "2027-06-14; duration on vacancy",
        "description": "Build sales and trading skills at the centre of global financial markets.",
        "requirements": "Penultimate-year internship profile; market interest, analytical thinking, collaboration and communication.",
        "language_work_rights": "Verify vacancy-specific language terms; Hong Kong work rights available.",
        "stages": "Online application and work-based assessments; interview stages as specified in candidate portal.",
        "documents": "CV; transcript if requested; work-authorization details; assessment preparation",
        "source": "https://www.hsbc.com/careers/students-and-graduates/find-a-programme?programme-type=internship-programme",
        "confidence": "Confirmed 2027 official listing",
        "notes": "Rolling-risk alert: submit well before 30 October.",
    },
    {
        "category": "Investment Banking & Finance",
        "company": "HSBC",
        "role": "Global Investment Research Internship",
        "season": "Summer 2027",
        "location": "Central, Hong Kong",
        "status": "Open",
        "open_date": d("2026-07-06"),
        "deadline": d("2026-10-30"),
        "deadline_tz": "Hong Kong time (confirm on form)",
        "rolling": "May close early",
        "placement": "2027-06-14; duration on vacancy",
        "description": "Research markets and trends to support investment and business decisions.",
        "requirements": "Penultimate-year internship profile; intellectual curiosity, research, analytical and communication skills.",
        "language_work_rights": "Verify vacancy-specific language terms; Hong Kong work rights available.",
        "stages": "Online application and work-based assessments; interview stages as specified in candidate portal.",
        "documents": "CV; transcript if requested; work-authorization details; assessment preparation",
        "source": "https://www.hsbc.com/careers/students-and-graduates/find-a-programme?programme-type=internship-programme",
        "confidence": "Confirmed 2027 official listing",
        "notes": "Strong match for a business/finance applicant interested in markets and company analysis.",
    },
    {
        "category": "Investment Banking & Finance",
        "company": "Goldman Sachs",
        "role": "2027 Summer Analyst Program - Asia Pacific",
        "season": "Summer 2027",
        "location": "Hong Kong subject to application availability",
        "status": "Open / verify location",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not published; rolling",
        "rolling": "Yes",
        "placement": "9-11 weeks",
        "description": "Immersive internship with training and real responsibilities across businesses including Investment Banking, FICC and Equities, Research, Asset Management, Operations and Risk.",
        "requirements": "Current penultimate-year Bachelor's or Master's student; apply as soon as ready because applications are reviewed rolling.",
        "language_work_rights": "Location-specific terms appear in each application; Hong Kong work rights available.",
        "stages": "Application and business-specific selection process.",
        "documents": "CV; academic information/transcript if requested; work-authorization details; division preferences",
        "source": "https://www.goldmansachs.com/careers/students/programs-and-internships/asia-pacific/summer-analyst",
        "confidence": "Confirmed 2027 official programme",
        "notes": "Check Apply Now immediately for live Hong Kong divisions; no single programme-wide deadline is published.",
    },
    {
        "category": "Investment Banking & Finance",
        "company": "Bank of America",
        "role": "Global Capital Markets Summer Analyst - 2027",
        "season": "Summer 2027",
        "location": "Hong Kong",
        "status": "Open",
        "open_date": None,
        "deadline": d("2026-09-30"),
        "deadline_tz": "Confirm on application",
        "rolling": "Apply early",
        "placement": "10 weeks",
        "description": "Research companies and markets, support transaction structuring and execution, and work with Investment Banking and Global Markets teams.",
        "requirements": "Penultimate-year undergraduate or master's student for 2028 conversion; strong academics, quantitative and analytical skills, communication, initiative and interest in investment banking.",
        "language_work_rights": "English required; another Asian language highly advantageous. Verify work authorization on application.",
        "stages": "Online application followed by programme-specific assessment and interviews.",
        "documents": "CV; academic record/transcript; work-authorization details; interview and technical preparation",
        "source": "https://careers.bankofamerica.com/en-us/students/job-detail/14375/global-capital-markets-summer-analyst-2027-hong-kong-hong-kong-hong-kong",
        "confidence": "Confirmed 2027 official role",
        "notes": "Direct fit for a 2028 graduate; deadline is employer-published.",
    },
    {
        "category": "Investment Banking & Finance",
        "company": "J.P. Morgan",
        "role": "Investment Banking Seasonal Internship",
        "season": "Winter / off-cycle watch",
        "location": "Hong Kong when displayed as open",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not published",
        "rolling": "Location-dependent",
        "placement": "3-6 months during spring, fall or winter",
        "description": "Participate in live transactions, valuation, research, pitch materials and deal execution with investment banking teams.",
        "requirements": "Analytical thinking, attention to detail, communication, business writing, research, quantitative ability and interest in financial services.",
        "language_work_rights": "Location-specific requirements apply; Hong Kong work rights available.",
        "stages": "Apply when Hong Kong appears in the official location selector; standard assessment and interviews follow.",
        "documents": "CV; transcript if requested; work-authorization details; technical interview preparation",
        "source": "https://careers.jpmorgan.com/US/en/students/programs/investment-banking-seasonal-analyst",
        "confidence": "Official programme; HK opening not confirmed",
        "notes": "Useful winter/off-cycle target. The official page only displays locations that are currently open.",
    },
    {
        "category": "Consulting",
        "company": "Deloitte China",
        "role": "Summer or Winter Internship - Consulting",
        "season": "Winter 2026-27 / Summer 2027 watch",
        "location": "Hong Kong subject to vacancy",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not yet published",
        "rolling": "Programme cycle",
        "placement": "2-4 months",
        "description": "Structured internship in Consulting, Audit or Tax with client exposure, networking, coaching and feedback.",
        "requirements": "Students from any discipline; generally penultimate year or above, or graduating in the next one to two years; strong interest in professional services.",
        "language_work_rights": "Vacancy-specific; applicant has Hong Kong work rights and relevant language capability.",
        "stages": "Application and selection process shown when role opens.",
        "documents": "CV; transcript; cover letter if requested; work-authorization details",
        "source": "https://www.deloitte.com/cn/en/careers/explore-your-fit/students/internship-program.html?icid=toggle_cn_en",
        "confidence": "Official recurring timeline; 2027 dates not confirmed",
        "notes": "Official guidance: winter recruitment normally Oct-Nov; summer recruitment normally Apr-May.",
    },
    {
        "category": "Consulting",
        "company": "KPMG China",
        "role": "Basecamp Internship - Advisory",
        "season": "Winter 2026-27 / Summer 2027 watch",
        "location": "Hong Kong and Macau stream",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not yet published",
        "rolling": "Multiple intakes",
        "placement": "Winter, summer, off-cycle or placement intake",
        "description": "Client experience and professional training in Advisory, Audit, Tax or Business Support, with internship opportunities across university years.",
        "requirements": "University student; strong academics, communication, industry interest, initiative and teamwork. Hong Kong office requires HKID or valid work permit.",
        "language_work_rights": "Permanent Hong Kong ID or valid work permit for Hong Kong office.",
        "stages": "Select intake, function and Chinese HK/MO location when applications open.",
        "documents": "CV; transcript; cover letter if requested; HKID/work-permit evidence",
        "source": "https://kpmg.com/cn/en/careers/campus/basecamp.html",
        "confidence": "Official programme; 2027 intake not published",
        "notes": "Monitor Basecamp for the next winter and summer intake tiles.",
    },
    {
        "category": "Consulting",
        "company": "Bain & Company",
        "role": "Associate Consultant Internship",
        "season": "Summer 2027 watch",
        "location": "Hong Kong subject to office opening",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Office-specific",
        "rolling": "Office-specific",
        "placement": "Typically 10 weeks",
        "description": "Work with consulting teams on strategic client challenges after intensive training in core consulting skills.",
        "requirements": "Office-specific eligibility; strong academics, analytical and creative thinking, teamwork and ability to thrive in an entrepreneurial environment.",
        "language_work_rights": "Office-specific; verify Hong Kong language and authorization terms.",
        "stages": "Application, office screening and case interviews.",
        "documents": "CV; educational background; work experience; test scores if applicable; office-specific items",
        "source": "https://www.bain.com/careers/work-with-us/internships-programs/",
        "confidence": "Official programme; HK 2027 date not confirmed",
        "notes": "Use the official office selector and apply only when Hong Kong is available.",
    },
    {
        "category": "Consulting",
        "company": "McKinsey & Company",
        "role": "Summer Business Analyst Internship",
        "season": "Summer 2027 watch",
        "location": "Hong Kong subject to job search",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Programme/office-specific",
        "rolling": "Varies",
        "placement": "Programme-specific",
        "description": "Hands-on problem solving across industries and functions as part of client-service teams.",
        "requirements": "Undergraduate pathway; strong problem solving, leadership, collaboration and academic performance. Exact eligibility is posting-specific.",
        "language_work_rights": "Hong Kong posting-specific; verify language and authorization terms.",
        "stages": "Application, problem-solving assessment where used, and interviews.",
        "documents": "CV; transcript/education details; work-authorization details; case interview preparation",
        "source": "https://www.mckinsey.com/careers/students/en",
        "confidence": "Official student pathway; HK 2027 date not confirmed",
        "notes": "Use McKinsey's deadline finder and job search for Hong Kong updates.",
    },
    {
        "category": "Supply Chain & Operations",
        "company": "Cathay Pacific",
        "role": "Summer Internship Programme - Business / Operations watch",
        "season": "Summer 2027 watch",
        "location": "Hong Kong",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not yet published",
        "rolling": "Vacancy-specific",
        "placement": "Prior cycle ran July-August",
        "description": "Paid project-based internships across functions; prior business roles included strategy, customer, digital and operational work.",
        "requirements": "Prior postings required full-time availability, strong English, interest in aviation and Hong Kong right of abode or acceptable student work documentation.",
        "language_work_rights": "English; right of abode or vacancy-approved documentation such as a No Objection Letter.",
        "stages": "Online application and function-specific interviews.",
        "documents": "CV; transcript if requested; HK work-right evidence; availability confirmation",
        "source": "https://careers.cathaypacific.com/our-teams/future-talent",
        "confidence": "Official pathway; timeline uses prior-cycle context",
        "notes": "Monitor from early 2027; do not treat the prior-cycle dates as the 2027 deadline.",
    },
    {
        "category": "Supply Chain & Operations",
        "company": "DHL",
        "role": "Hong Kong Student Internship Opportunities",
        "season": "Winter / Summer 2027 watch",
        "location": "Hong Kong",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Vacancy-specific",
        "rolling": "Current opportunities",
        "placement": "Vacancy-specific",
        "description": "Student opportunities offering real-world exposure to logistics and supply-chain operations; roles vary by business and project need.",
        "requirements": "Enrolled student; opportunity-specific academic and availability requirements; supply-chain interest and analytical/team skills are advantageous.",
        "language_work_rights": "Hong Kong vacancy-specific; likely English and Chinese for local-facing roles.",
        "stages": "Apply through current opportunities; recruitment process varies by role.",
        "documents": "CV; transcript if requested; work-authorization details; availability",
        "source": "https://careers.dhl.com/apac/en/students-graduates",
        "confidence": "Official student pathway; HK 2027 role not confirmed",
        "notes": "The Hong Kong management-trainee page is graduate-focused; use student vacancies for internship eligibility.",
    },
    {
        "category": "General Business Programs",
        "company": "Swire",
        "role": "Swire Summer Internship Programme 2027",
        "season": "Summer 2027",
        "location": "Hong Kong for Hong Kong citizens",
        "status": "Upcoming",
        "open_date": d("2026-12-28"),
        "deadline": d("2027-01-31"),
        "deadline_tz": "23:59 HKT (UTC+8)",
        "rolling": "No",
        "placement": "2027-07-02 to 2027-08-20",
        "description": "Own an eight-week business project, develop new skills and pitch a business idea to senior management, with support from a programme buddy and line manager.",
        "requirements": "Penultimate-year undergraduate or postgraduate, graduating by July 2028; commercial acumen, initiative, learning agility, leadership potential, teamwork and strong English. Chinese is advantageous.",
        "language_work_rights": "Excellent English; Putonghua or Cantonese advantageous. Hong Kong citizens assigned to Hong Kong projects.",
        "stages": "Online application; video interview in February; virtual group case in March; final interviews in March; offers by mid-April.",
        "documents": "CV; academic details/transcript if requested; work-right/citizenship details; video and case preparation",
        "source": "https://careers.swire.com/en/careers/swire-summer-internship-programme",
        "confidence": "Confirmed 2027 official programme",
        "notes": "Excellent broad-business option with a defined application and selection timeline.",
    },
    {
        "category": "General Business Programs",
        "company": "HKEX",
        "role": "Summer Internship Programme 2027",
        "season": "Summer 2027",
        "location": "Hong Kong",
        "status": "Upcoming",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not yet published",
        "rolling": "Programme cycle",
        "placement": "10 weeks, June-August 2027",
        "description": "Direct project responsibility at the centre of global markets across functions including Finance, Strategy, Markets, Operations, Listing, Risk, Sustainability and HR.",
        "requirements": "Expected to target penultimate-year students; prior cycle required strong academics, commercial awareness, problem solving, adaptability, teamwork and excellent English. Confirm 2027 criteria when posted.",
        "language_work_rights": "Confirm in the 2027 posting; applicant profile is strongly aligned.",
        "stages": "Applications officially scheduled to open in September 2026; later stages will be published with the vacancy.",
        "documents": "CV; transcript; work-authorization details; division preferences; interview preparation",
        "source": "https://www.hkexgroup.com/About-HKEX/Careers-at-HKEX/Early-Careers?sc_lang=en",
        "confidence": "Official 2027 opening month; exact date pending",
        "notes": "Set an opening alert for 1 September 2026 and verify weekly until the application appears.",
    },
    {
        "category": "General Business Programs",
        "company": "Securities and Futures Commission",
        "role": "Summer and Winter Internship Programme",
        "season": "Winter 2026-27 / Summer 2027 watch",
        "location": "Hong Kong",
        "status": "Monitor",
        "open_date": d("2026-09-01"),
        "deadline": None,
        "deadline_tz": "Winter deadline not published",
        "rolling": "Programme cycle",
        "placement": "Winter usually Dec-Jan; summer six weeks in Jul-Aug",
        "description": "Support SFC divisions with research, reports, document management, systems updates and other operational or drafting work.",
        "requirements": "Undergraduate or postgraduate in fields including Law, Accounting, Economics, Finance, Management or IT; analytical, problem-solving and interpersonal skills; integrity; English and Chinese; Microsoft Office.",
        "language_work_rights": "Good spoken and written English and Chinese; local eligibility terms should be checked on the application.",
        "stages": "Winter application starts from September; summer application timing to be confirmed for 2027.",
        "documents": "CV; transcript; work-authorization details; availability; supporting documents requested on form",
        "source": "https://hksfc.org/TC/Career/Why-the-SFC/Join-us-as-an-Intern/Summer-and-Winter-Internship-Programme?sc_lang=en",
        "confidence": "Official recurring timeline; exact 2027 deadlines pending",
        "notes": "High language/profile fit. Check weekly from September for the winter application.",
    },
    {
        "category": "General Business Programs",
        "company": "Hong Kong Government",
        "role": "Administrative Service Summer Internship Programme",
        "season": "Summer 2027 watch",
        "location": "Hong Kong",
        "status": "Monitor",
        "open_date": None,
        "deadline": None,
        "deadline_tz": "Not yet published",
        "rolling": "Annual programme",
        "placement": "Prior cycle: two months or more between May and September",
        "description": "Public-service internship designed to build exposure to Administrative Service work and community-focused leadership.",
        "requirements": "Prior cycle: Hong Kong permanent resident; full-time tertiary student; leadership, teamwork, community involvement, strong academics; bilingual Chinese/English and trilingual Cantonese/Putonghua/English.",
        "language_work_rights": "Strong match for trilingual applicant; confirm 2027 residency and study-year rules.",
        "stages": "Local students usually apply through their tertiary institution; overseas students follow the official direct process.",
        "documents": "Institution nomination/application; CV or form; academic information; photo; residency evidence",
        "source": "https://www.ao-recruitment.gov.hk/english/summer_intern.html",
        "confidence": "Official prior-cycle reference; 2027 details pending",
        "notes": "Start monitoring in January 2027; prior cycle closed 27 February 2026.",
    },
]


def next_action(item):
    if item["status"] == "Open":
        if item["rolling"] in {"Yes", "May close early", "Apply early"}:
            return "Apply as soon as materials are ready"
        return "Prepare and submit application"
    if item["open_date"]:
        return f"Set opening alert for {item['open_date'].isoformat()}"
    return "Monitor official page in weekly reminder"


def deadline_alert(item):
    if item["status"] == "Open" and item["rolling"] in {"Yes", "May close early", "Apply early"}:
        return "OPEN NOW - rolling/early-close risk"
    if item["deadline"]:
        days = (item["deadline"] - TODAY).days
        if days < 0:
            return "Deadline passed"
        if days <= 7:
            return "7-day alert"
        if days <= 14:
            return "14-day alert"
        if days <= 30:
            return "30-day alert"
        return f"{days} days remaining"
    if item["open_date"] and item["open_date"] > TODAY:
        return f"Opens in {(item['open_date'] - TODAY).days} days"
    return "No confirmed deadline - monitor"


headers = [
    "ID", "Category", "Company", "Program / Role", "Season", "Location", "Application Status",
    "Opening Date", "Closing Date", "Deadline Time Zone", "Rolling?", "Placement Timeline",
    "Deadline Alert", "Next Action", "Job Description", "Requirements", "Language / Work Rights",
    "Recruitment Stages", "Document Checklist", "Official Application / Source", "Date Confidence", "Notes",
]


wb = Workbook()
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_opp = wb.create_sheet("Opportunities")
ws_pipe = wb.create_sheet("Application Pipeline")
ws_docs = wb.create_sheet("Document Checklist")
ws_companies = wb.create_sheet("Company Watchlist")
ws_readme = wb.create_sheet("Read Me")


def title_band(ws, title, subtitle=None):
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color=GRAY)


title_band(ws_dash, "Hong Kong 2027 Internship Tracker", "Profile: 2028 business/finance graduate | HK work rights | English, Cantonese and Mandarin")
ws_dash["A4"] = "Category selector"
ws_dash["B4"] = "All"
category_validation = DataValidation(type="list", formula1='"All,Consulting,Investment Banking & Finance,Supply Chain & Operations,General Business Programs"')
ws_dash.add_data_validation(category_validation)
category_validation.add(ws_dash["B4"])

metrics = [
    ("A6", "Tracked opportunities", '=IF(B4="All",COUNTA(Opportunities!C2:C200),COUNTIF(Opportunities!B2:B200,B4))'),
    ("C6", "Open now", '=IF(B4="All",COUNTIF(Opportunities!G2:G200,"Open*"),COUNTIFS(Opportunities!B2:B200,B4,Opportunities!G2:G200,"Open*"))'),
    ("E6", "Upcoming", '=IF(B4="All",COUNTIF(Opportunities!G2:G200,"Upcoming"),COUNTIFS(Opportunities!B2:B200,B4,Opportunities!G2:G200,"Upcoming"))'),
    ("G6", "Needs monitoring", '=IF(B4="All",COUNTIF(Opportunities!G2:G200,"Monitor"),COUNTIFS(Opportunities!B2:B200,B4,Opportunities!G2:G200,"Monitor"))'),
]
for cell, label, formula in metrics:
    ws_dash[cell] = label
    ws_dash[cell].font = Font(bold=True, color=GRAY)
    value_cell = ws_dash.cell(row=ws_dash[cell].row + 1, column=ws_dash[cell].column)
    value_cell.value = formula
    value_cell.font = Font(size=22, bold=True, color=NAVY)

ws_dash["A10"] = "Urgent actions"
ws_dash["A10"].font = Font(size=14, bold=True, color=NAVY)
urgent_headers = ["Company", "Role", "Category", "Alert", "Deadline", "Next Action"]
for col, header in enumerate(urgent_headers, 1):
    c = ws_dash.cell(row=11, column=col, value=header)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.font = Font(bold=True, color=WHITE)

urgent = [x for x in opportunities if x["status"].startswith("Open") or x["status"] == "Upcoming"]
urgent.sort(key=lambda x: (x["deadline"] or date.max, x["open_date"] or date.max))
for row_idx, item in enumerate(urgent[:10], 12):
    vals = [item["company"], item["role"], item["category"], deadline_alert(item), item["deadline"], next_action(item)]
    for col, val in enumerate(vals, 1):
        ws_dash.cell(row=row_idx, column=col, value=val)

ws_dash["A24"] = "Reminder cadence"
ws_dash["A24"].font = Font(size=14, bold=True, color=NAVY)
ws_dash["A25"] = "Weekly scan"
ws_dash["B25"] = "Monday, 9:00 AM Hong Kong time"
ws_dash["A26"] = "Deadline alerts"
ws_dash["B26"] = "Opening date; 30 days when applicable; 14 days; 7 days"
ws_dash["A27"] = "Evidence rule"
ws_dash["B27"] = "Official employer sources first; inferred dates must remain labelled"

for col, width in {"A": 25, "B": 32, "C": 25, "D": 30, "E": 17, "F": 38, "G": 22, "H": 22}.items():
    ws_dash.column_dimensions[col].width = width
ws_dash.freeze_panes = "A11"


for col_idx, header in enumerate(headers, 1):
    cell = ws_opp.cell(row=1, column=col_idx, value=header)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(wrap_text=True, vertical="center")

for idx, item in enumerate(opportunities, 1):
    row = [
        f"HK27-{idx:03d}", item["category"], item["company"], item["role"], item["season"], item["location"],
        item["status"], item["open_date"], item["deadline"], item["deadline_tz"], item["rolling"], item["placement"],
        deadline_alert(item), next_action(item), item["description"], item["requirements"], item["language_work_rights"],
        item["stages"], item["documents"], item["source"], item["confidence"], item["notes"],
    ]
    for col_idx, value in enumerate(row, 1):
        cell = ws_opp.cell(row=idx + 1, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    link_cell = ws_opp.cell(row=idx + 1, column=20)
    link_cell.hyperlink = item["source"]
    link_cell.style = "Hyperlink"

for col in (8, 9):
    for row in range(2, len(opportunities) + 2):
        ws_opp.cell(row=row, column=col).number_format = "yyyy-mm-dd"

table = Table(displayName="OpportunityTable", ref=f"A1:V{len(opportunities)+1}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
ws_opp.add_table(table)
ws_opp.freeze_panes = "A2"
ws_opp.auto_filter.ref = f"A1:V{len(opportunities)+1}"
widths = [12, 30, 22, 42, 25, 28, 20, 14, 14, 24, 18, 28, 32, 38, 55, 60, 48, 55, 55, 58, 38, 55]
for idx, width in enumerate(widths, 1):
    ws_opp.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width
ws_opp.row_dimensions[1].height = 44

red_fill = PatternFill("solid", fgColor=LIGHT_RED)
yellow_fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
ws_opp.conditional_formatting.add(f"M2:M{len(opportunities)+1}", FormulaRule(formula=['ISNUMBER(SEARCH("7-day",M2))'], fill=red_fill))
ws_opp.conditional_formatting.add(f"M2:M{len(opportunities)+1}", FormulaRule(formula=['OR(ISNUMBER(SEARCH("14-day",M2)),ISNUMBER(SEARCH("30-day",M2)))'], fill=yellow_fill))
ws_opp.conditional_formatting.add(f"G2:G{len(opportunities)+1}", FormulaRule(formula=['ISNUMBER(SEARCH("Open",G2))'], fill=green_fill))


pipeline_headers = ["Opportunity ID", "Company", "Program / Role", "Category", "Pipeline Status", "Priority", "Owner", "Last Updated", "Next Task", "Next Task Due", "Notes"]
for col, header in enumerate(pipeline_headers, 1):
    c = ws_pipe.cell(row=1, column=col, value=header)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(bold=True, color=WHITE)
for idx, item in enumerate(opportunities, 2):
    values = [f"HK27-{idx-1:03d}", item["company"], item["role"], item["category"], "Interested", "Medium", "Me", TODAY, next_action(item), item["deadline"] or item["open_date"], ""]
    for col, value in enumerate(values, 1):
        ws_pipe.cell(row=idx, column=col, value=value)
status_validation = DataValidation(type="list", formula1='"Interested,Preparing,Applied,Testing,Interviewing,Offer,Rejected,Closed"')
priority_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
ws_pipe.add_data_validation(status_validation)
ws_pipe.add_data_validation(priority_validation)
status_validation.add(f"E2:E500")
priority_validation.add(f"F2:F500")
pipe_table = Table(displayName="PipelineTable", ref=f"A1:K{len(opportunities)+1}")
pipe_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
ws_pipe.add_table(pipe_table)
ws_pipe.freeze_panes = "A2"
for col, width in enumerate([15, 22, 42, 30, 18, 12, 12, 15, 42, 15, 45], 1):
    ws_pipe.column_dimensions[chr(64 + col)].width = width
for row in range(2, len(opportunities) + 2):
    ws_pipe.cell(row=row, column=8).number_format = "yyyy-mm-dd"
    ws_pipe.cell(row=row, column=10).number_format = "yyyy-mm-dd"


doc_rows = [
    ["Master CV", "Core", "Preparing", "Tailor achievements for consulting, finance, operations and general business variants", ""],
    ["Consulting CV", "Consulting", "Not started", "Emphasize leadership, impact, problem solving and quantified outcomes", ""],
    ["Finance CV", "Investment Banking & Finance", "Not started", "Emphasize finance interest, analytical work, transactions, markets and technical skills", ""],
    ["Operations CV", "Supply Chain & Operations", "Not started", "Emphasize process improvement, analytics, teamwork and execution", ""],
    ["General Business CV", "General Business Programs", "Not started", "Emphasize commercial judgment, leadership, adaptability and cross-functional projects", ""],
    ["Master cover-letter bank", "All", "Not started", "Prepare modular motivation, company-fit and experience paragraphs", ""],
    ["Official transcript", "All", "Not started", "Request a current PDF and confirm grading scale", ""],
    ["HK work-right evidence", "All", "Ready", "Keep HKID or relevant authorization evidence available where requested", ""],
    ["English/Cantonese/Mandarin profile", "All", "Ready", "Use consistent proficiency wording across applications", ""],
    ["Online assessment preparation", "All", "Not started", "Practice numerical, verbal, situational judgment and work-style assessments", ""],
    ["Consulting case preparation", "Consulting", "Not started", "Build case framework practice schedule and fit-interview story bank", ""],
    ["Finance technical preparation", "Investment Banking & Finance", "Not started", "Accounting, valuation, markets, commercial awareness and deal discussion", ""],
]
doc_headers = ["Document / Task", "Category", "Readiness", "Action", "Target Date"]
for col, header in enumerate(doc_headers, 1):
    c = ws_docs.cell(row=1, column=col, value=header)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(bold=True, color=WHITE)
for row_idx, values in enumerate(doc_rows, 2):
    for col_idx, value in enumerate(values, 1):
        ws_docs.cell(row=row_idx, column=col_idx, value=value)
ready_validation = DataValidation(type="list", formula1='"Not started,Preparing,Ready,Needs update"')
ws_docs.add_data_validation(ready_validation)
ready_validation.add("C2:C200")
doc_table = Table(displayName="DocumentTable", ref=f"A1:E{len(doc_rows)+1}")
doc_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws_docs.add_table(doc_table)
ws_docs.freeze_panes = "A2"
for col, width in enumerate([32, 32, 18, 75, 18], 1):
    ws_docs.column_dimensions[chr(64 + col)].width = width


companies = {}
for item in opportunities:
    key = (item["category"], item["company"])
    companies.setdefault(key, {"source": item["source"], "status": item["status"], "notes": item["notes"]})
company_headers = ["Category", "Company", "Current Tracker Status", "Official Careers / Programme Link", "Monitoring Note"]
for col, header in enumerate(company_headers, 1):
    c = ws_companies.cell(row=1, column=col, value=header)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(bold=True, color=WHITE)
for row_idx, ((category, company), info) in enumerate(sorted(companies.items()), 2):
    vals = [category, company, info["status"], info["source"], info["notes"]]
    for col_idx, value in enumerate(vals, 1):
        ws_companies.cell(row=row_idx, column=col_idx, value=value)
    ws_companies.cell(row=row_idx, column=4).hyperlink = info["source"]
    ws_companies.cell(row=row_idx, column=4).style = "Hyperlink"
company_table = Table(displayName="CompanyTable", ref=f"A1:E{len(companies)+1}")
company_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_companies.add_table(company_table)
ws_companies.freeze_panes = "A2"
for col, width in enumerate([32, 25, 25, 65, 80], 1):
    ws_companies.column_dimensions[chr(64 + col)].width = width


title_band(ws_readme, "How to Use This Tracker", "Updated 15 July 2026; dates must be rechecked before submission")
instructions = [
    ("1. Choose a category", "Use Dashboard cell B4 or the filters on Opportunities to switch among Consulting, Investment Banking & Finance, Supply Chain & Operations and General Business Programs."),
    ("2. Act on confirmed openings", "Start with the Dashboard urgent-actions section. Rolling or early-close roles should be submitted as soon as materials are ready."),
    ("3. Treat confidence labels seriously", "Confirmed means the employer has published 2027 information. Monitoring or prior-cycle entries are planning aids, not confirmed deadlines."),
    ("4. Update pipeline status", "Use Interested, Preparing, Applied, Testing, Interviewing, Offer, Rejected or Closed in the Application Pipeline sheet."),
    ("5. Prepare documents", "Use Document Checklist to maintain category-specific CVs, transcript, work-right evidence and assessment preparation."),
    ("6. Weekly reminder", "The Skywork automation runs every Monday at 9:00 AM Hong Kong time and should report new openings, deadline changes and required actions."),
    ("7. Verify before applying", "Open the official source link and reconfirm title, location, deadline time zone, eligibility and required documents immediately before submission."),
]
for idx, (heading, body) in enumerate(instructions, 4):
    ws_readme[f"A{idx}"] = heading
    ws_readme[f"A{idx}"].font = Font(bold=True, color=NAVY)
    ws_readme[f"B{idx}"] = body
    ws_readme[f"B{idx}"].alignment = Alignment(wrap_text=True, vertical="top")
ws_readme.column_dimensions["A"].width = 28
ws_readme.column_dimensions["B"].width = 105


thin = Side(style="thin", color=GRID)
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    if ws.max_row > 1 and ws.max_column > 1:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(bottom=thin)

wb.save(OUTPUT)
print(OUTPUT.resolve())
